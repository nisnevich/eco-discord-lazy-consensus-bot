import logging
import io
import discord
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from bot.config.const import (
    HELP_MESSAGE_NON_AUTHORIZED_USER,
    HELP_MESSAGE_AUTHORIZED_USER,
    RESPONSIBLE_MENTION,
    HELP_COMMAND_NAME,
    DEFAULT_LOG_LEVEL,
    EXPORT_COMMAND_NAME,
    EMPTY_ANALYTICS_VALUE,
)
from bot.config.logging_config import log_handler, console_handler
from bot.utils.discord_utils import get_discord_client, get_message, get_user_by_id_or_mention
from bot.utils.validation import validate_roles
from bot.utils.db_utils import DBUtil
from bot.utils.formatting_utils import get_amount_to_print
from bot.config.schemas import ProposalHistory
from bot.config.const import ProposalResult, VOTING_CHANNEL_ID

logger = logging.getLogger(__name__)
logger.setLevel(DEFAULT_LOG_LEVEL)
logger.addHandler(log_handler)
logger.addHandler(console_handler)

client = get_discord_client()


@client.command(name=HELP_COMMAND_NAME)
async def help(ctx):
    try:
        # Remove the help request message
        await ctx.message.delete()
        # Reply to a non-authorized user
        if not await validate_roles(ctx.message.author):
            await ctx.author.send(HELP_MESSAGE_NON_AUTHORIZED_USER)
            return
        # Reply to an authorized user
        await ctx.author.send(HELP_MESSAGE_AUTHORIZED_USER)
    except Exception as e:
        try:
            # Try replying in Discord
            await ctx.message.reply(
                f"An unexpected error occurred when sending help. cc {RESPONSIBLE_MENTION}"
            )
        except Exception as e:
            logger.critical("Unable to reply in the chat that a critical error has occurred.")

        logger.critical(
            "Unexpected error in %s while sending help, channel=%s, message=%s, user=%s",
            __name__,
            ctx.message.channel.id,
            ctx.message.id,
            ctx.message.author.mention,
            exc_info=True,
        )


def export_docx(accepted_proposals):
    # Create a new Excel workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Proposal History"

    # Define column names and widths
    columns = [
        {"header": "Discord link", "width": 20},
        {"header": "When completed (UTC time)", "width": 25},
        {"header": "Author", "width": 20},
        {"header": "Has grant", "width": 10},
        {"header": "Given to", "width": 25},
        {"header": "Amount", "width": 15},
        {"header": "Description", "width": 40},
    ]

    # Write the column names to the worksheet and set column widths
    for col_num, column in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)
        column_header = column["header"]
        column_width = column["width"]
        worksheet.column_dimensions[column_letter].width = column_width
        worksheet.cell(row=1, column=col_num, value=column_header).font = Font(bold=True)

    # Loop over each accepted proposal and add a row to the worksheet
    for row_num, proposal in enumerate(accepted_proposals, 2):
        # Add a hyperlink to the Discord link column
        discord_link = proposal.discord_link
        worksheet.cell(row=row_num, column=1).value = discord_link
        worksheet.cell(row=row_num, column=1).hyperlink = discord_link

        # Add data to the remaining columns
        worksheet.cell(
            row=row_num, column=2, value=proposal.closed_at.strftime("%Y-%m-%d %H:%M:%S")
        )
        worksheet.cell(row=row_num, column=3, value=str(proposal.author))
        worksheet.cell(row=row_num, column=4, value=str(not proposal.is_grantless))
        worksheet.cell(
            row=row_num,
            column=5,
            value=str(proposal.mention) if proposal.mention is not None else EMPTY_ANALYTICS_VALUE,
        )
        worksheet.cell(
            row=row_num,
            column=6,
            value=str(get_amount_to_print(proposal.amount))
            if proposal.amount is not None
            else EMPTY_ANALYTICS_VALUE,
        )
        worksheet.cell(row=row_num, column=7, value=str(proposal.description))

    # Create a table for the worksheet
    table_range = f"A1:{get_column_letter(len(columns))}{len(accepted_proposals)+1}"
    table = Table(displayName="ProposalHistory", ref=table_range)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    # Save the Excel workbook to a temporary file
    temp_file = io.BytesIO()
    workbook.save(temp_file)
    temp_file.seek(0)

    return temp_file, "proposal_history.xlsx"


def export_csv(accepted_proposals):
    file = io.StringIO()
    writer = csv.DictWriter(
        file,
        fieldnames=[
            "When completed (UTC time)",
            "Author",
            "Has grant",
            "Given to",
            "Amount",
            "Description",
            "Voting URL",
        ],
    )
    writer.writeheader()

    for proposal in accepted_proposals:
        logger.debug("Exporting %s", proposal)
        writer.writerow(
            {
                "When completed (UTC time)": proposal.closed_at.strftime("%Y-%m-%d %H:%M:%S"),
                "Author": proposal.author,
                "Has grant": not proposal.is_grantless,
                "Given to": proposal.mention
                if proposal.mention is not None
                else EMPTY_ANALYTICS_VALUE,
                "Amount": get_amount_to_print(proposal.amount)
                if proposal.amount is not None
                else EMPTY_ANALYTICS_VALUE,
                "Description": proposal.description,
                "Voting URL": proposal.voting_message_url,
            }
        )
    file.seek(0)

    return file, "proposal_history.csv"


async def export(ctx, is_csv):
    try:
        # Reply to a non-authorized user
        if not await validate_roles(ctx.message.author):
            # Adding greetings and "cancelled" reactions
            await original_message.add_reaction(REACTION_ON_BOT_MENTION)
            await original_message.add_reaction(CANCEL_EMOJI_UNICODE)
            # Sending response in DM
            await ctx.author.send(HELP_MESSAGE_NON_AUTHORIZED_USER)
            return

        # Remove the message requesting the analytics
        await ctx.message.delete()

        accepted_proposals = (
            DBUtil.session_history.query(ProposalHistory)
            .filter(ProposalHistory.result == ProposalResult.ACCEPTED.value)
            .all()
        )
        if len(accepted_proposals) == 0:
            await ctx.author.send("No proposals were accepted yet.")
            return

        if is_csv:
            document, filename = export_csv(accepted_proposals)
        else:
            document, filename = export_docx(accepted_proposals)

        await ctx.author.send(file=discord.File(document, filename=filename))

    except Exception as e:
        try:
            # Try replying in Discord
            await ctx.message.reply(
                f"An unexpected error occurred when exporting analytical data. cc {RESPONSIBLE_MENTION}"
            )
        except Exception as e:
            logger.critical("Unable to reply in the chat that a critical error has occurred.")

        logger.critical(
            "Unexpected error in %s while exporting analytical data, channel=%s, message=%s, user=%s",
            __name__,
            ctx.message.channel.id,
            ctx.message.id,
            ctx.message.author.mention,
            exc_info=True,
        )


@client.command(name=EXPORT_COMMAND_NAME)
async def export_command(ctx):
    export(ctx, False)


@client.command(name=EXPORT_CSV_COMMAND_NAME)
async def export_command(ctx):
    export(ctx, True)
